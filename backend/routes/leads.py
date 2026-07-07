from datetime import datetime
from flask import Blueprint, request, jsonify, current_app, send_file
from flask_jwt_extended import jwt_required, get_jwt
from utils.helpers import admin_required, employee_or_admin, serialize_doc, serialize_list
from bson.objectid import ObjectId
import io
import openpyxl

leads_bp = Blueprint('leads', __name__)


@leads_bp.route('', methods=['GET'])
@jwt_required()
def get_leads():
    current_user = get_jwt()
    db = current_app.config['db']

    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 50))
    search = request.args.get('search', '')
    status_filter = request.args.get('status', '')

    query = {}
    if current_user['role'] == 'employee':
        query['assigned_to'] = current_user['full_name']
    if search:
        query['$or'] = [
            {'name': {'$regex': search, '$options': 'i'}},
            {'phone': {'$regex': search, '$options': 'i'}},
            {'email': {'$regex': search, '$options': 'i'}},
            {'school_name': {'$regex': search, '$options': 'i'}},
            {'company': {'$regex': search, '$options': 'i'}},
            {'city': {'$regex': search, '$options': 'i'}},
            {'lead_source': {'$regex': search, '$options': 'i'}},
        ]
    if status_filter:
        query['status'] = status_filter

    total = db.leads.count_documents(query)
    leads = list(db.leads.find(query, {
        '_id': 1, 'name': 1, 'phone': 1, 'whatsapp': 1, 'email': 1,
        'status': 1, 'assigned_to': 1, 'lead_source': 1, 'city': 1,
        'company': 1, 'organization': 1, 'created_at': 1,
    }).sort('created_at', 1)
                 .skip((page - 1) * limit)
                 .limit(limit))

    employees = None
    if current_user['role'] == 'admin':
        employees = list(db.users.find(
            {'role': 'employee', 'is_active': True},
            {'full_name': 1}
        ))

    return jsonify({
        'leads': serialize_list(leads),
        'total': total,
        'page': page,
        'pages': (total + limit - 1) // limit,
        'employees': employees,
    })


@leads_bp.route('', methods=['POST'])
@employee_or_admin
def create_lead():
    current_user = get_jwt()
    data = request.get_json()
    db = current_app.config['db']

    exclude = {'lead_id', 'id', '_id', 'created_at', 'updated_at', 'created_by', 'contacted_count', 'last_contacted'}
    lead = {k: v for k, v in data.items() if k not in exclude}

    lead.setdefault('name', '')
    lead.setdefault('phone', '')
    lead.setdefault('status', 'New')
    lead.setdefault('assigned_to', current_user['full_name'])
    if 'whatsapp' not in lead and lead.get('phone'):
        lead['whatsapp'] = lead['phone']
    lead['contacted_count'] = 0
    lead['last_contacted'] = None
    lead['created_by'] = current_user['full_name']
    lead['created_at'] = datetime.utcnow()
    lead['updated_at'] = datetime.utcnow()

    result = db.leads.insert_one(lead)
    lead['_id'] = str(result.inserted_id)

    return jsonify({'message': 'Lead created', 'lead': lead}), 201


@leads_bp.route('/<lead_id>', methods=['GET'])
@jwt_required()
def get_lead(lead_id):
    db = current_app.config['db']
    lead = db.leads.find_one({'_id': ObjectId(lead_id)})
    if not lead:
        return jsonify({'error': 'Lead not found'}), 404
    return jsonify({'lead': serialize_doc(lead)})


@leads_bp.route('/<lead_id>', methods=['PUT'])
@jwt_required()
def update_lead(lead_id):
    current_user = get_jwt()
    data = request.get_json()
    db = current_app.config['db']

    lead = db.leads.find_one({'_id': ObjectId(lead_id)})
    if not lead:
        return jsonify({'error': 'Lead not found'}), 404

    if current_user['role'] == 'employee' and lead.get('assigned_to') != current_user['full_name']:
        return jsonify({'error': 'Not authorized to update this lead'}), 403

    exclude = {'lead_id', 'id', '_id', 'created_at', 'updated_at', 'created_by', 'contacted_count', 'last_contacted'}
    update = {k: v for k, v in data.items() if k not in exclude}

    if 'contacted' in data and data['contacted']:
        update['contacted_count'] = (lead.get('contacted_count', 0) + 1)
        update['last_contacted'] = datetime.utcnow()

    update['updated_at'] = datetime.utcnow()

    db.leads.update_one({'_id': ObjectId(lead_id)}, {'$set': update})
    lead = db.leads.find_one({'_id': ObjectId(lead_id)})

    return jsonify({'message': 'Lead updated', 'lead': serialize_doc(lead)})


@leads_bp.route('/<lead_id>', methods=['DELETE'])
@admin_required
def delete_lead(lead_id):
    db = current_app.config['db']
    db.leads.delete_one({'_id': ObjectId(lead_id)})
    return jsonify({'message': 'Lead deleted'})


@leads_bp.route('/import', methods=['POST'])
@admin_required
def import_leads():
    db = current_app.config['db']
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    wb = openpyxl.load_workbook(io.BytesIO(file.read()))
    ws = wb.active

    imported = 0
    duplicates = 0
    skipped = 0
    errors = []

    rows_iter = ws.iter_rows(values_only=True)
    first_row = next(rows_iter, None)
    has_header = first_row and any(str(h).strip().lower() in ('name', 'lead', 'customer', 'contact', 'school name') for h in first_row if h)

    if has_header:
        headers = [str(h).strip().lower().replace(' ', '_').replace('/', '_').replace('-', '_') if h else '' for h in first_row]
    else:
        headers = ['name', 'phone', 'email', 'whatsapp', 'address', 'status', 'assigned_to']

    start_row = 2 if has_header else 1
    for i, row in enumerate(rows_iter if has_header else ws.iter_rows(min_row=1, values_only=True), start=start_row):
        try:
            cells = list(row) if row else []
            if not any(c is not None for c in cells):
                skipped += 1
                continue

            lead = {'contacted_count': 0, 'last_contacted': None, 'created_by': 'import', 'created_at': datetime.utcnow(), 'updated_at': datetime.utcnow()}
            for ci, val in enumerate(cells):
                if ci < len(headers) and headers[ci] and val is not None:
                    key = headers[ci]
                    lead[key] = str(val).strip()

            name = lead.get('name', '') or ''
            phone = lead.get('phone', '') or ''
            if not name and not phone:
                skipped += 1
                continue

            if phone and db.leads.find_one({'phone': phone}):
                duplicates += 1
                continue

            lead.setdefault('status', 'New')
            lead.setdefault('whatsapp', phone)

            db.leads.insert_one(lead)
            imported += 1
        except Exception as e:
            errors.append(f'Row {i}: {str(e)}')

    return jsonify({
        'message': f'Imported {imported} leads (skipped {skipped}, duplicates {duplicates})',
        'imported': imported,
        'duplicates': duplicates,
        'skipped': skipped,
        'errors': errors,
    })


@leads_bp.route('/export', methods=['GET'])
@admin_required
def export_leads():
    db = current_app.config['db']
    leads = list(db.leads.find().sort('created_at', -1))

    all_keys = ['name', 'phone', 'email', 'whatsapp', 'status', 'assigned_to', 'created_by']
    for l in leads:
        for k in l:
            if k not in all_keys and k not in ('_id', 'contacted_count', 'last_contacted', 'created_at', 'updated_at'):
                all_keys.append(k)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Leads'
    ws.append(all_keys)

    for l in leads:
        row = []
        for k in all_keys:
            v = l.get(k, '')
            if isinstance(v, datetime):
                v = str(v)
            row.append(v)
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='leads.xlsx'
    )


@leads_bp.route('/bulk-delete', methods=['POST'])
@admin_required
def bulk_delete_leads():
    db = current_app.config['db']
    data = request.get_json()
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'error': 'No IDs provided'}), 400

    obj_ids = [ObjectId(id) for id in ids]
    result = db.leads.delete_many({'_id': {'$in': obj_ids}})
    return jsonify({'message': f'Deleted {result.deleted_count} leads', 'count': result.deleted_count})


@leads_bp.route('/delete-all', methods=['DELETE'])
@admin_required
def delete_all_leads():
    db = current_app.config['db']
    result = db.leads.delete_many({})
    return jsonify({'message': f'Deleted {result.deleted_count} leads', 'count': result.deleted_count})
