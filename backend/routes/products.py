from datetime import datetime
from flask import Blueprint, request, jsonify, current_app, send_file
from flask_jwt_extended import jwt_required
from utils.helpers import admin_required, employee_or_admin, serialize_doc, serialize_list
from bson.objectid import ObjectId
import io
import openpyxl

products_bp = Blueprint('products', __name__)


@products_bp.route('', methods=['GET'])
@jwt_required()
def get_products():
    db = current_app.config['db']
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 50))
    search = request.args.get('search', '')
    category = request.args.get('category', '')

    query = {'is_active': True}
    if search:
        query['$or'] = [
            {'name': {'$regex': search, '$options': 'i'}},
            {'sku': {'$regex': search, '$options': 'i'}},
        ]
    if category:
        query['category'] = category

    total = db.products.count_documents(query)
    products = list(db.products.find(query)
                    .sort('name', 1)
                    .skip((page - 1) * limit)
                    .limit(limit))

    categories = db.products.distinct('category', {'is_active': True})

    return jsonify({
        'products': serialize_list(products),
        'total': total,
        'page': page,
        'pages': (total + limit - 1) // limit,
        'categories': categories,
    })


@products_bp.route('/<product_id>', methods=['GET'])
@jwt_required()
def get_product(product_id):
    db = current_app.config['db']
    product = db.products.find_one({'_id': ObjectId(product_id)})
    if not product:
        return jsonify({'error': 'Product not found'}), 404
    return jsonify({'product': serialize_doc(product)})


@products_bp.route('', methods=['POST'])
@admin_required
def create_product():
    data = request.get_json()
    db = current_app.config['db']

    product = {
        'name': data.get('name', ''),
        'sku': data.get('sku', ''),
        'category': data.get('category', 'General'),
        'price': float(data.get('price', 0)),
        'cost_price': float(data.get('cost_price', 0)),
        'selling_price': float(data.get('selling_price', 0)),
        'mrp': float(data.get('mrp', 0)),
        'stock': int(data.get('stock', 0)),
        'description': data.get('description', ''),
        'is_active': True,
        'created_at': datetime.utcnow(),
        'updated_at': datetime.utcnow(),
    }

    result = db.products.insert_one(product)
    product['_id'] = str(result.inserted_id)

    return jsonify({'message': 'Product created', 'product': product}), 201


@products_bp.route('/<product_id>', methods=['PUT'])
@admin_required
def update_product(product_id):
    data = request.get_json()
    db = current_app.config['db']

    update = {}
    for field in ['name', 'sku', 'category', 'description']:
        if field in data:
            update[field] = data[field]
    for field in ['price', 'cost_price', 'selling_price', 'mrp']:
        if field in data:
            update[field] = float(data[field])
    if 'stock' in data:
        update['stock'] = int(data['stock'])
    if 'is_active' in data:
        update['is_active'] = data['is_active']

    update['updated_at'] = datetime.utcnow()

    db.products.update_one({'_id': ObjectId(product_id)}, {'$set': update})
    product = db.products.find_one({'_id': ObjectId(product_id)})

    return jsonify({'message': 'Product updated', 'product': serialize_doc(product)})


@products_bp.route('/<product_id>', methods=['DELETE'])
@admin_required
def delete_product(product_id):
    db = current_app.config['db']
    db.products.delete_one({'_id': ObjectId(product_id)})
    return jsonify({'message': 'Product deleted'})


@products_bp.route('/import', methods=['POST'])
@admin_required
def import_products():
    db = current_app.config['db']
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    wb = openpyxl.load_workbook(io.BytesIO(file.read()))
    ws = wb.active

    imported = 0
    duplicates = 0
    skipped = 0
    errors = []

    # Try to detect header row
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    has_header = first_row and any(str(h).strip().lower() in ('name', 'product', 'sku', 'product name') for h in first_row if h)

    for i, row in enumerate(ws.iter_rows(min_row=2 if has_header else 1, values_only=True), start=2 if has_header else 1):
        try:
            name = ''
            sku = ''
            category = 'General'
            price = 0
            cost_price = 0
            selling_price = 0
            mrp = 0
            stock = 0

            # Flexible column detection: try to use all available cells
            cells = [c for c in row if c is not None]
            if not cells:
                continue

            name = str(cells[0]).strip() if len(cells) > 0 else ''
            if len(cells) > 1:
                sku = str(cells[1]).strip()
            if len(cells) > 2:
                category = str(cells[2]).strip()
            if len(cells) > 3 and cells[3]:
                try: price = float(cells[3])
                except: pass
            if len(cells) > 4 and cells[4]:
                try: cost_price = float(cells[4])
                except: pass
            if len(cells) > 5 and cells[5]:
                try: selling_price = float(cells[5])
                except: pass
            if len(cells) > 6 and cells[6]:
                try: mrp = float(cells[6])
                except: pass
            if len(cells) > 7 and cells[7]:
                try: stock = int(float(cells[7]))
                except: pass

            if not name:
                skipped += 1
                continue

            if sku and db.products.find_one({'sku': sku}):
                duplicates += 1
                continue

            # Auto-set selling_price if not provided
            if not selling_price and price:
                selling_price = price

            product = {
                'name': name,
                'sku': sku,
                'category': category,
                'price': price,
                'cost_price': cost_price,
                'selling_price': selling_price or price,
                'mrp': mrp or selling_price or price,
                'stock': stock,
                'description': '',
                'is_active': True,
                'created_at': datetime.utcnow(),
                'updated_at': datetime.utcnow(),
            }
            db.products.insert_one(product)
            imported += 1
        except Exception as e:
            errors.append(f'Row {i}: {str(e)}')

    return jsonify({
        'message': f'Imported {imported} products (skipped {skipped}, duplicates {duplicates})',
        'imported': imported,
        'duplicates': duplicates,
        'skipped': skipped,
        'errors': errors,
    })


@products_bp.route('/export', methods=['GET'])
@admin_required
def export_products():
    db = current_app.config['db']
    products = list(db.products.find({'is_active': True}).sort('name', 1))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Products'
    ws.append(['Name', 'SKU', 'Category', 'Price', 'Cost Price', 'Selling Price', 'MRP', 'Stock'])

    for p in products:
        ws.append([
            p.get('name', ''),
            p.get('sku', ''),
            p.get('category', ''),
            p.get('price', 0),
            p.get('cost_price', 0),
            p.get('selling_price', 0),
            p.get('mrp', 0),
            p.get('stock', 0),
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='products.xlsx'
    )


@products_bp.route('/seed', methods=['POST'])
@admin_required
def seed_products():
    db = current_app.config['db']
    count = db.products.count_documents({})
    if count > 0:
        return jsonify({'message': f'Products already exist ({count})'}), 400

    products = [
        # Wooden Toys
        {'name': 'Wooden Building Blocks (50 pcs)', 'sku': 'WT-001', 'category': 'Wooden Toys', 'price': 899, 'cost_price': 500, 'selling_price': 899, 'mrp': 1299, 'stock': 50},
        {'name': 'Wooden Puzzle - Animals', 'sku': 'WT-002', 'category': 'Wooden Toys', 'price': 349, 'cost_price': 180, 'selling_price': 349, 'mrp': 499, 'stock': 40},
        {'name': 'Wooden Shape Sorter', 'sku': 'WT-003', 'category': 'Wooden Toys', 'price': 449, 'cost_price': 250, 'selling_price': 449, 'mrp': 649, 'stock': 35},
        {'name': 'Wooden Stacking Rings', 'sku': 'WT-004', 'category': 'Wooden Toys', 'price': 399, 'cost_price': 200, 'selling_price': 399, 'mrp': 599, 'stock': 30},
        {'name': 'Wooden Abacus', 'sku': 'WT-005', 'category': 'Wooden Toys', 'price': 599, 'cost_price': 320, 'selling_price': 599, 'mrp': 849, 'stock': 25},
        # Early Practical Life
        {'name': 'Montessori Dressing Frame - Button', 'sku': 'PL-001', 'category': 'Practical Life', 'price': 799, 'cost_price': 450, 'selling_price': 799, 'mrp': 1099, 'stock': 20},
        {'name': 'Pouring Set with Jug', 'sku': 'PL-002', 'category': 'Practical Life', 'price': 649, 'cost_price': 350, 'selling_price': 649, 'mrp': 899, 'stock': 15},
        {'name': 'Scooping & Transfer Set', 'sku': 'PL-003', 'category': 'Practical Life', 'price': 549, 'cost_price': 300, 'selling_price': 549, 'mrp': 749, 'stock': 20},
        {'name': 'Montessori Sweeping Set', 'sku': 'PL-004', 'category': 'Practical Life', 'price': 699, 'cost_price': 380, 'selling_price': 699, 'mrp': 999, 'stock': 18},
        # Sensorial Materials
        {'name': 'Pink Tower', 'sku': 'SM-001', 'category': 'Sensorial', 'price': 2499, 'cost_price': 1500, 'selling_price': 2499, 'mrp': 3499, 'stock': 10},
        {'name': 'Brown Stair', 'sku': 'SM-002', 'category': 'Sensorial', 'price': 2199, 'cost_price': 1300, 'selling_price': 2199, 'mrp': 2999, 'stock': 8},
        {'name': 'Knobbed Cylinders', 'sku': 'SM-003', 'category': 'Sensorial', 'price': 1899, 'cost_price': 1100, 'selling_price': 1899, 'mrp': 2699, 'stock': 12},
        {'name': 'Color Tablets Box 1', 'sku': 'SM-004', 'category': 'Sensorial', 'price': 599, 'cost_price': 350, 'selling_price': 599, 'mrp': 849, 'stock': 20},
        {'name': 'Sound Boxes', 'sku': 'SM-005', 'category': 'Sensorial', 'price': 899, 'cost_price': 500, 'selling_price': 899, 'mrp': 1299, 'stock': 15},
        {'name': 'Touch Boards', 'sku': 'SM-006', 'category': 'Sensorial', 'price': 749, 'cost_price': 420, 'selling_price': 749, 'mrp': 1049, 'stock': 15},
        # Language Materials
        {'name': 'Sandpaper Letters (Lowercase)', 'sku': 'LM-001', 'category': 'Language', 'price': 1299, 'cost_price': 750, 'selling_price': 1299, 'mrp': 1799, 'stock': 20},
        {'name': 'Moveable Alphabet (Print)', 'sku': 'LM-002', 'category': 'Language', 'price': 1799, 'cost_price': 1000, 'selling_price': 1799, 'mrp': 2499, 'stock': 12},
        {'name': 'Metal Insets (10 shapes)', 'sku': 'LM-003', 'category': 'Language', 'price': 1499, 'cost_price': 850, 'selling_price': 1499, 'mrp': 2099, 'stock': 10},
        {'name': 'Phonics Object Box', 'sku': 'LM-004', 'category': 'Language', 'price': 999, 'cost_price': 550, 'selling_price': 999, 'mrp': 1399, 'stock': 15},
        # Math Materials
        {'name': 'Number Rods', 'sku': 'MM-001', 'category': 'Math', 'price': 2199, 'cost_price': 1300, 'selling_price': 2199, 'mrp': 2999, 'stock': 8},
        {'name': 'Spindle Box', 'sku': 'MM-002', 'category': 'Math', 'price': 1299, 'cost_price': 750, 'selling_price': 1299, 'mrp': 1799, 'stock': 10},
        {'name': 'Golden Bead Material (Complete)', 'sku': 'MM-003', 'category': 'Math', 'price': 4999, 'cost_price': 3000, 'selling_price': 4999, 'mrp': 6999, 'stock': 5},
        {'name': 'Teen Board with Beads', 'sku': 'MM-004', 'category': 'Math', 'price': 1599, 'cost_price': 900, 'selling_price': 1599, 'mrp': 2199, 'stock': 10},
        {'name': 'Hundred Board', 'sku': 'MM-005', 'category': 'Math', 'price': 999, 'cost_price': 550, 'selling_price': 999, 'mrp': 1399, 'stock': 15},
        {'name': 'Stamp Game', 'sku': 'MM-006', 'category': 'Math', 'price': 799, 'cost_price': 450, 'selling_price': 799, 'mrp': 1099, 'stock': 18},
        # Geography
        {'name': 'Wooden World Map Puzzle', 'sku': 'GG-001', 'category': 'Geography', 'price': 1799, 'cost_price': 1000, 'selling_price': 1799, 'mrp': 2499, 'stock': 10},
        {'name': 'Continent Globe', 'sku': 'GG-002', 'category': 'Geography', 'price': 1499, 'cost_price': 850, 'selling_price': 1499, 'mrp': 2099, 'stock': 8},
        {'name': 'Land & Water Forms Tray', 'sku': 'GG-003', 'category': 'Geography', 'price': 1199, 'cost_price': 680, 'selling_price': 1199, 'mrp': 1699, 'stock': 12},
        # Botany & Zoology
        {'name': 'Wooden Plant Puzzle', 'sku': 'BZ-001', 'category': 'Botany & Zoology', 'price': 699, 'cost_price': 400, 'selling_price': 699, 'mrp': 999, 'stock': 15},
        {'name': 'Insect Puzzle Set', 'sku': 'BZ-002', 'category': 'Botany & Zoology', 'price': 849, 'cost_price': 480, 'selling_price': 849, 'mrp': 1199, 'stock': 12},
        {'name': 'Animal Classification Cards', 'sku': 'BZ-003', 'category': 'Botany & Zoology', 'price': 549, 'cost_price': 300, 'selling_price': 549, 'mrp': 799, 'stock': 20},
        {'name': 'Life Cycle Puzzles (Set of 4)', 'sku': 'BZ-004', 'category': 'Botany & Zoology', 'price': 1299, 'cost_price': 750, 'selling_price': 1299, 'mrp': 1799, 'stock': 8},
        # Furniture
        {'name': 'Montessori Table (Kids, 60cm)', 'sku': 'FN-001', 'category': 'Furniture', 'price': 4999, 'cost_price': 2800, 'selling_price': 4999, 'mrp': 6999, 'stock': 10},
        {'name': 'Montessori Chair (Kids)', 'sku': 'FN-002', 'category': 'Furniture', 'price': 2499, 'cost_price': 1400, 'selling_price': 2499, 'mrp': 3499, 'stock': 20},
        {'name': 'Open Shelf Unit (3-tier)', 'sku': 'FN-003', 'category': 'Furniture', 'price': 8499, 'cost_price': 5000, 'selling_price': 8499, 'mrp': 11999, 'stock': 5},
        {'name': 'Work Mat (60x80cm)', 'sku': 'FN-004', 'category': 'Furniture', 'price': 299, 'cost_price': 150, 'selling_price': 299, 'mrp': 449, 'stock': 50},
        {'name': 'Floor Cushion', 'sku': 'FN-005', 'category': 'Furniture', 'price': 449, 'cost_price': 250, 'selling_price': 449, 'mrp': 649, 'stock': 30},
        # Art & Music
        {'name': 'Montessori Bell Set', 'sku': 'AM-001', 'category': 'Art & Music', 'price': 1999, 'cost_price': 1200, 'selling_price': 1999, 'mrp': 2799, 'stock': 8},
        {'name': 'Easel Board (Double Sided)', 'sku': 'AM-002', 'category': 'Art & Music', 'price': 2999, 'cost_price': 1700, 'selling_price': 2999, 'mrp': 4199, 'stock': 6},
        {'name': 'Rhythm Instrument Set', 'sku': 'AM-003', 'category': 'Art & Music', 'price': 899, 'cost_price': 500, 'selling_price': 899, 'mrp': 1299, 'stock': 15},
    ]

    for p in products:
        p['is_active'] = True
        p['description'] = ''
        p['created_at'] = datetime.utcnow()
        p['updated_at'] = datetime.utcnow()

    db.products.insert_many(products)
    return jsonify({'message': f'Seeded {len(products)} products', 'count': len(products)})
